import streamlit as st
from openpyxl import load_workbook, Workbook
import re
import io

# --- Page setup ---
st.set_page_config(page_title="Excel Sheet Combiner", layout="centered")

st.title("📊 Excel Sheet Combiner (Preserves Formatting & Formulas)")
st.write("Combine sheets from multiple Excel files while keeping ALL formatting and formulas.")

uploaded_files = st.file_uploader(
    "📁 Upload Excel Files",
    type=["xlsx"],
    accept_multiple_files=True
)

sheet_to_extract = st.text_input(
    "🔎 Sheet name to extract (leave empty to extract ALL sheets):",
    value=""
)

st.markdown("---")

# ----------- Helper function: clean filename ----------
def clean_filename(filename: str) -> str:
    # Example: "20260211_4467 South Acton ESG.xlsx" → "South Acton ESG"
    filename = filename.replace(".xlsx", "")
    cleaned = re.sub(r"^\d+[_\-]*\d*\s*", "", filename).strip()
    return cleaned


# ----------- Main Logic -------------
if st.button("🚀 Combine Excel Files"):
    if not uploaded_files:
        st.error("⚠ Please upload at least one file.")
        st.stop()

    progress = st.progress(0)
    status = st.empty()

    output_wb = Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)  # Remove blank sheet

    total = len(uploaded_files)

    for index, uploaded in enumerate(uploaded_files):
        status.text(f"Processing **{uploaded.name}** ...")

        # Load workbook with formulas preserved
        wb = load_workbook(uploaded, data_only=False)

        cleaned_name = clean_filename(uploaded.name)

        # If specific sheet requested
        if sheet_to_extract:
            if sheet_to_extract not in wb.sheetnames:
                st.warning(f"⚠ Sheet '{sheet_to_extract}' not found in {uploaded.name}")
            else:
                source_sheet = wb[sheet_to_extract]
                new_sheet_name = f"{cleaned_name} - {sheet_to_extract}"
                new_sheet = output_wb.create_sheet(new_sheet_name)

                # Copy cell values + styles
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet[cell.coordinate]
                        new_cell.value = cell.value

                        if cell.has_style:
                            if cell.font:
                                new_cell.font = cell.font.copy()
                            if cell.border:
                                new_cell.border = cell.border.copy()
                            if cell.fill:
                                new_cell.fill = cell.fill.copy()
                            if cell.number_format:
                                new_cell.number_format = cell.number_format
                            if cell.protection:
                                new_cell.protection = cell.protection.copy()
                            if cell.alignment:
                                new_cell.alignment = cell.alignment.copy()

                # Copy merged cells
                for merged_range in source_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_range))

                # Copy column widths & row heights
                for col_letter, dim in source_sheet.column_dimensions.items():
                    new_sheet.column_dimensions[col_letter].width = dim.width

                for row_idx, dim in source_sheet.row_dimensions.items():
                    new_sheet.row_dimensions[row_idx].height = dim.height

        else:
            # Extract ALL sheets
            for sh in wb.sheetnames:
                source_sheet = wb[sh]
                new_sheet_name = f"{cleaned_name} - {sh}"
                new_sheet = output_wb.create_sheet(new_sheet_name)

                # Copy cell values + styles
                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet[cell.coordinate]
                        new_cell.value = cell.value

                        if cell.has_style:
                            if cell.font:
                                new_cell.font = cell.font.copy()
                            if cell.border:
                                new_cell.border = cell.border.copy()
                            if cell.fill:
                                new_cell.fill = cell.fill.copy()
                            if cell.number_format:
                                new_cell.number_format = cell.number_format
                            if cell.protection:
                                new_cell.protection = cell.protection.copy()
                            if cell.alignment:
                                new_cell.alignment = cell.alignment.copy()

                # Copy merged cells
                for merged_range in source_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_range))

                # Copy dimensions
                for col_letter, dim in source_sheet.column_dimensions.items():
                    new_sheet.column_dimensions[col_letter].width = dim.width

                for row_idx, dim in source_sheet.row_dimensions.items():
                    new_sheet.row_dimensions[row_idx].height = dim.height

        progress.progress((index + 1) / total)

    # Save output
    output_path = "combined.xlsx"
    output_wb.save(output_path)

    with open(output_path, "rb") as f:
        st.download_button(
            "⬇ Download Combined Excel File",
            f,
            file_name="combined.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.success("🎉 Done! Sheets combined with FULL formatting + formulas preserved.")
